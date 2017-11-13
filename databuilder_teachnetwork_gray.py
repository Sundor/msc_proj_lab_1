import os
from PIL import Image
import time
from random import shuffle
import pickle
import datetime
import sys
import numpy as np
import math
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import h5py

from keras.preprocessing.image import ImageDataGenerator
from keras.models import Sequential, load_model, save_model
from keras.layers import Conv2D, MaxPooling2D
from keras.layers import Activation, Dropout, Flatten, Dense
from keras.callbacks import EarlyStopping, ModelCheckpoint, ReduceLROnPlateau
from keras.optimizers import rmsprop

# Dictionary containing all paths to the processed images
path_d = {
    'train_pics': "H:\\Downloads\\data_object_image_2\\training\\image_2",
    'train_labels': "H:\\Downloads\\data_object_label_2\\training\\label_2",
    'Car': "H:\\Databases\\data_object_cars\\processed",
    'Van': "H:\\Databases\\data_object_vans\\processed",
    'Truck': "H:\\Databases\\data_object_trucks\\processed",
    'Pedestrian': "H:\\Databases\\data_object_pedestrians\\processed",
    'Person_sitting': "H:\\Databases\\data_object_sittingpeople\\processed",
    'Cyclist': "H:\\Databases\\data_object_cyclists\\processed",
    'Tram': "H:\\Databases\\data_object_trams",
    'Misc': "H:\\Databases\\data_object_misc",
    'DontCare': "H:\\Databases\\data_object_dontcare",
    'Background': "H:\\Databases\\data_object_backgrounds",
    'MainDatabase': "H:\\Databases\\processed_database\\grayscale",
    'Model': "H:\\Databases\\model_weights",
    'Log': "H:\\Databases\\model_weights\\results.xlsx"
}

# Dictionary containing column numbers in log .xlsx file
log_cols = {
    'model_name':   1,
    'start':        2,
    'time_elapsed':     3,

    'labels':       4,
    'num_data':     5,
    'train_split':  6,
    'valid_split':  7,
    'batch_size':   8,
    'net_shape':    9,

    'epochs':       10,
    'loss':         11,
    'acc':          12,
    'val_loss':     13,
    'val_acc':      14
}

# Load pictures from directory, create training, validation and test databases
def create_databases(path_d, labels, train_split, valid_split):

    main_db = []

    start_time = time.time()
    num_curr_file = 0

    # read total number of files
    num_files = 0
    for cur_label,label in enumerate(labels):
        files = os.listdir(path_d[label])
        num_files += len(files)
    print("Processing "+ str(num_files) +" pictures.")

    # get every picture from
    for cur_label,label in enumerate(labels):
        num_labels = len(labels)

        files = os.listdir(path_d[label])
        for file_in_label, file in enumerate(files):
            num_curr_file += 1
            if (file_in_label>40000):
                break           # !!!!!!!!!!!!!!!!
            if file[-3:] == "png":
                pic = Image.open(path_d[label] + "\\" + file).convert('L')
                # input, pixels of the picture
                x = np.array(pic.getdata())
                y = np.zeros(num_labels, dtype=np.uint8)
                # output, correct label
                y[cur_label] = 1
                main_db.append((x, y)) # or [x, y] ????
                #print
                percent = int(round(((num_curr_file+1) / num_files) * 100))
                sys.stdout.write("\rReading pictures: %d%% - Label: %s - Size of database: %s" %(percent, label, str(sys.getsizeof(main_db)/1024/1024)))
                pic.close()
    print("\nShuffle database.")
    shuffle(main_db)

    train_db, valid_db, test_db = split_databases(main_db, train_split, valid_split)

    dir = path_d['MainDatabase']+"\\maindata_" + datetime.datetime.now().strftime("%Y%m%d_%H%M%S") + "_" + str(len(labels))
    os.makedirs(dir)

    save_path = dir + "\\train.txt"
    savef = open(save_path, "wb")
    pickle.dump(train_db, savef)
    savef.close()
    print("Database saved to: " + save_path)

    save_path = dir + "\\valid.txt"
    savef = open(save_path, "wb")
    pickle.dump(valid_db, savef)
    savef.close()
    print("Database saved to: " + save_path)

    save_path = dir + "\\test.txt"
    savef = open(save_path, "wb")
    pickle.dump(test_db, savef)
    savef.close()
    print("Database saved to: " + save_path)

    print("Runtime: " + str(int(time.time() - start_time)) + " sec.")

    return train_db, valid_db


def load_databases(path):
    curr_path = path + "\\train.txt"
    loadf = open(curr_path, "rb")
    train_db = pickle.load(loadf)
    loadf.close()

    curr_path = path + "\\valid.txt"
    loadf = open(curr_path, "rb")
    valid_db = pickle.load(loadf)
    loadf.close()

    return train_db, valid_db


def split_databases(main_db, train_split, valid_split):
    nb_samples = len(main_db)
    nb_train_samples = math.floor(nb_samples * train_split)
    nb_valid_samples = math.floor(nb_samples * valid_split)
    nb_test_samples = nb_samples - nb_valid_samples - nb_train_samples

    train_database = main_db[: nb_train_samples]
    valid_database = main_db[nb_train_samples: nb_train_samples + nb_valid_samples]
    test_database = main_db[nb_train_samples + nb_valid_samples:]

    train_x = np.array([row[0].reshape(64, 64, 1) for row in train_database])
    train_y = np.array([row[1] for row in train_database])
    train_xy = (train_x, train_y)

    valid_x = np.array([row[0].reshape(64, 64, 1) for row in valid_database])
    valid_y = np.array([row[1] for row in valid_database])
    valid_xy = (valid_x, valid_y)

    test_x = np.array([row[0].reshape(64, 64, 1) for row in test_database])
    test_y = np.array([row[1] for row in test_database])
    test_xy = (test_x, test_y)

    print("Databases created:")
    print("Train samples: " + str(nb_train_samples))
    print("Validation samples: " + str(nb_valid_samples))
    print("Test samples: " + str(nb_test_samples))

    return train_xy, valid_xy, test_xy


def create_model():
    model = Sequential()

    model.add(Conv2D(32, (5, 5), input_shape=picture_array_shape))
    model.add(Activation('relu'))
    model.add(MaxPooling2D(pool_size=(2, 2)))

    model.add(Conv2D(32, (3, 3)))
    model.add(Activation('relu'))
    model.add(MaxPooling2D(pool_size=(2, 2)))

    model.add(Conv2D(64, (3, 3)))
    model.add(Activation('relu'))
    model.add(MaxPooling2D(pool_size=(2, 2)))

    model.add(Flatten())
    model.add(Dense(64))
    model.add(Activation('relu'))
    model.add(Dropout(0.5))
    model.add(Dense(num_labels))
    model.add(Activation('sigmoid'))

    model.compile(loss='binary_crossentropy',
              optimizer='adam',
              metrics=['accuracy'])

    return model


def train_model(model, train_xy, valid_xy, epochs, batch_size, earlystopping):
    train_x = train_xy[0]
    train_y = train_xy[1]

    valid_x = valid_xy[0]
    valid_y = valid_xy[1]

    nb_train_samples = len(train_x)
    nb_valid_samples = len(valid_x)

    callbacks = [EarlyStopping(monitor='val_loss', min_delta=0.001, patience=earlystopping, verbose=0),
                 ModelCheckpoint(weights_save_path, monitor='val_loss', verbose=0, save_best_only=True,
                                 save_weights_only=True),
                 ReduceLROnPlateau(monitor='val_loss', factor=0.5, patience=5, min_lr=0.001)
                 ]

    # this is the augmentation configuration we will use for training
    train_datagen = ImageDataGenerator(shear_range=0.2,
                                       zoom_range=0.2,
                                       horizontal_flip=True)

    # this is the augmentation configuration we will use for testing:
    # only rescaling
    valid_datagen = ImageDataGenerator()

    train_generator = train_datagen.flow(
        train_x,
        train_y,
        batch_size=batch_size,
        shuffle=True,
        seed=36)

    validation_generator = valid_datagen.flow(
        valid_x,
        valid_y,
        batch_size=batch_size,
        shuffle=True,
        seed=36)

    history = model.fit_generator(
        train_generator,
        steps_per_epoch=nb_train_samples // batch_size,
        epochs=epochs,
        validation_data=validation_generator,
        validation_steps=nb_valid_samples // batch_size,
        callbacks=callbacks,
        verbose=2)

    model.load_weights(weights_save_path)  # load weights last saved
    model.save(model_save_path)
    print("Model saved as: " + model_save_path)

    return model, history


def log_results(paths, model, history, epochs):
    time_elapsed = (end_time - start_time).total_seconds()
    num_epochs = epochs

    labels_str = ""
    for label in used_labels:
        labels_str += label + ", "

    divider = " >>> "
    net_shape = "Layers: " + str(len(model.layers)) + "\n"
    for layer in model.layers:

        if "activation" in layer.name:
            net_shape += "   " + layer.output.name + "\n"

        elif "conv2d" in layer.name:
            net_shape += layer.name + divider
            net_shape += "input:" + str(layer.input_shape)
            net_shape += ", kernel:" + str(layer.kernel_size) + "@" + str(layer.filters)
            net_shape += ", strides:" + str(layer.strides)
            net_shape += ", padding:" + str(layer.padding)
            net_shape += "\n"

        elif "max_pooling2d" in layer.name:
            net_shape += layer.name + divider
            net_shape += "input:" + str(layer.input_shape)
            net_shape += ", pool:" + str(layer.pool_size)
            net_shape += ", strides:" + str(layer.strides)
            net_shape += ", padding:" + str(layer.padding)
            net_shape += "\n"

        elif "flatten" in layer.name:
            net_shape += layer.name + divider
            net_shape += "output:" + str(layer.output_shape)
            net_shape += "\n"

        elif "dense" in layer.name:
            net_shape += layer.name + divider
            net_shape += "input:" + str(layer.input_shape)
            net_shape += ", nodes:" + str(layer.units)
            net_shape += "\n"

        elif "dropout" in layer.name:
            net_shape += "   " + layer.name + divider
            net_shape += "rate:" + str(layer.rate)
            net_shape += "\n"

        else:
            net_shape += layer.name + "\n"

    #print(net_shape)

    # open log file
    wb = load_workbook(paths['Log'])
    sheet = wb.worksheets[0]

    # Read prevoius results, if training was continued
    if cfg_create_model == False:
        for row in sheet.rows:
            if row[log_cols['model_name'] - 1].value == model_load_path:
                time_elapsed += row[log_cols['time_elapsed'] - 1].value
                num_epochs += row[log_cols['epochs'] - 1].value
                break

    new_row = sheet.max_row + 1

    # identify model
    sheet.cell(row=new_row, column=log_cols['model_name']).value = model_save_path
    sheet.cell(row=new_row, column=log_cols['start']).value = start_time.strftime("%Y. %m. %d. %H:%M:%S")
    sheet.cell(row=new_row, column=log_cols['time_elapsed']).value = time_elapsed
    # parameters
    sheet.cell(row=new_row, column=log_cols['labels']).value = labels_str
    sheet.cell(row=new_row, column=log_cols['num_data']).value = num_data
    sheet.cell(row=new_row, column=log_cols['train_split']).value = cfg_train_split
    sheet.cell(row=new_row, column=log_cols['valid_split']).value = cfg_valid_split
    sheet.cell(row=new_row, column=log_cols['batch_size']).value = cfg_batch_size
    sheet.cell(row=new_row, column=log_cols['net_shape']).value = net_shape
    sheet.cell(row=new_row, column=log_cols['net_shape']).alignment = Alignment(wrap_text=True)

    # results
    sheet.cell(row=new_row, column=log_cols['epochs']).value = num_epochs
    sheet.cell(row=new_row, column=log_cols['loss']).value = history.history['loss'][-1]
    sheet.cell(row=new_row, column=log_cols['acc']).value = history.history['acc'][-1]
    sheet.cell(row=new_row, column=log_cols['val_loss']).value = history.history['val_loss'][-1]
    sheet.cell(row=new_row, column=log_cols['val_acc']).value = history.history['val_acc'][-1]

    # format
    sheet.cell(row=new_row, column=log_cols['net_shape']).alignment = Alignment(wrap_text=True)
    sheet.cell(row=new_row, column=log_cols['labels']).alignment = Alignment(wrap_text=True)

    # close wb
    wb.save(paths['Log'])
    wb.close()
    print("Model logged.")


# save time and date of start
start_time_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
start_time = datetime.datetime.now()

#############
# CONFIGURED PARAMETERS

cfg_create_db = False # if true, gather pictures from data_object folders, and pickle a new database
cfg_create_model = False # if false, load a previous model and train that

load_path = path_d['MainDatabase'] + "\\maindata_20171113_175240_6"
model_load_path = path_d['Model'] + "\\model_20171113_175054.hdf5"
#weights_load_path = path_d['Model'] + "\\weights_20171028_224634.hdf5"
used_labels = ['Car', 'Background', 'Pedestrian', 'Van', 'Truck', 'Cyclist']

picture_shape = (64, 64)
picture_array_shape = (64, 64, 1)

cfg_train_split = 0.7
cfg_valid_split = 0.15
cfg_epochs = 10
cfg_batch_size = 32
cfg_earlystopping = 30

#############
# CALCULATED PARAMETERS

num_labels = len(used_labels)
model_save_path = path_d['Model'] + "\\model_" + start_time_str + ".hdf5"
weights_save_path = path_d['Model'] + "\\weights_" + start_time_str + ".hdf5"

################
# GET DATABASE

if (cfg_create_db):
    train_xy, valid_xy = create_databases(path_d, used_labels, cfg_train_split, cfg_valid_split)
else:
    train_xy, valid_xy = load_databases(load_path)

num_data = len(train_xy[0]) + len(valid_xy[0])
########################
# CREATE MODEL
if cfg_epochs > 0:
    if (cfg_create_model):
        model = create_model()
        model, history = train_model(model, train_xy, valid_xy, cfg_epochs, cfg_batch_size, cfg_earlystopping)

    else:
        model = load_model(model_load_path)
        model, history = train_model(model, train_xy, valid_xy, cfg_epochs, cfg_batch_size, cfg_earlystopping)

    end_time = datetime.datetime.now()

    log_results(path_d, model, history, cfg_epochs)


