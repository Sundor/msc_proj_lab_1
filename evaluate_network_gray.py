import os
import sys
from keras import models
import matplotlib.pyplot as plt
import pickle
import math
import numpy as np
from PIL import Image
from openpyxl import load_workbook
import glob

path_d = {'train_pics': "H:\\Downloads\\data_object_image_2\\training\\image_2",
          'train_labels': "H:\\Downloads\\data_object_label_2\\training\\label_2",
          'Car': "H:\\Databases\\data_object_cars\\processed",
          'Van': "H:\\Databases\\data_object_vans\\processed",
          'Truck': "H:\\Databases\\data_object_trucks\\processed",
          'Pedestrian': "H:\\Databases\\data_object_pedestrians\\processed",
          'Person_sitting': "H:\\Databases\\data_object_sittingpeople\\processed",
          'Cyclist': "H:\\Databases\\data_object_cyclists\\processed",
          'Tram': "H:\\Databases\\data_object_trams",
          'Misc': "H:\\Databases\\data_object_misc",
          'DontCare': "H:\\Databases\\data_object_dontcare",
          'Background': "H:\\Databases\\data_object_backgrounds",
          'MainDatabase': "H:\\Databases\\processed_database\\grayscale",
          'Model': "H:\\Databases\\model_weights",
          'Outliers': "H:\\Dropbox\\BME-MSC\\2017-2018-1\\Önlab1\\outliers.xlsx"}


def load_databases(path):
    loadf = open(path + "\\test.txt", "rb")
    test_db = pickle.load(loadf)
    loadf.close()

    return test_db


def plot_conf_mat(conf_mat, labels, normal=False):
    fig, ax = plt.subplots()
    plt.gcf().subplots_adjust(top=0.75)

    ticks = np.arange(len(labels))
    plt.xticks(ticks, labels, rotation=45)
    ax.xaxis.tick_top()
    ax.xaxis.set_label_position('top')
    plt.yticks(ticks, labels)
    plt.ylabel('Truth')
    plt.xlabel('Predictions')


    # from: http://scikit-learn.org/stable/auto_examples/model_selection/plot_confusion_matrix.html#sphx-glr-auto-examples-model-selection-plot-confusion-matrix-py
    fmt = '.2f' if normal else 'd'
    thresh = conf_mat.max() / 2.
    # truth labels
    for i in range(conf_mat.shape[1]):
        sum_truths_per_label = sum(conf_mat[i])
        # predicted labels
        for j in range(conf_mat.shape[0]):
            num = conf_mat[i, j]
            if normal:
                num /= sum_truths_per_label
            plt.text(j, i, format(num, fmt),
                    horizontalalignment="center",
                    color="white" if conf_mat[i, j] < thresh else "black")

    plt.imshow(conf_mat, cmap='plasma')
    plt.show(block=True)


def findOutliers(model, labels, paths):
    # open log file
    wb = load_workbook(paths['Outliers'])
    sheet = wb.worksheets[0]

    for n_label in range(len(labels)):
        label = labels[n_label]
        y = np.zeros(num_labels, dtype=np.uint8)
        y[n_label] = 1
        truth_index = n_label

        files = os.listdir(path_d[label])

        for i,file in enumerate(files):
            sys.stdout.write("\rFinding outliers in: " + label + ", " + str(i))
            if file[-3:] == "png":
                pic = Image.open(path_d[label] + "\\" + file)
                x = np.array(pic.getdata()).reshape(1,64,64,3)

                prediction = model.predict(x, batch_size=1, verbose=0)
                prediction = prediction[0]
                order = np.argsort(prediction)
                preds = [[labels[j], prediction[j]] for j in reversed(order)]
                pred_index = np.argmax(prediction)

                if preds[0][1]>0.5 and pred_index != truth_index:
                    new_row = sheet.max_row + 1

                    pred_str = ""
                    for pred in preds:
                        pred_str += "%.2f %s," % (pred[1], pred[0])

                    sheet.cell(row=new_row, column=1).value = file
                    sheet.cell(row=new_row, column=2).value = "Prediction: " + pred_str
                    sheet.cell(row=new_row, column=3).value = "Truth: %s\n" % (used_labels[truth_index])

                    #pic_array = test_x
                    #img = Image.fromarray(pic_array.astype(np.uint8), 'RGB')
                    #img.show()
                    #plt.pause(5)

    # close wb
    wb.save(paths['Outliers'])
    wb.close()
    print("\nModel logged.")

#############
# CONFIGURED PARAMETERS
cfg_print_debug = True
cfg_find_outliers = False
cfg_use_latest_model = True

load_path = path_d['MainDatabase'] + "\\maindata_20171113_175240_6"
#cfg_model_load_path = path_d['Model'] + "\\model_20171109_155748.hdf5"
used_labels = ['Car', 'Background', 'Pedestrian', 'Van', 'Truck', 'Cyclist']

picture_shape = (64, 64)
picture_array_shape = (64, 64, 1)

train_split = 0.7
valid_split = 0.15
batch_size = 20

num_prediction = 20



#############
# CALCULATED PARAMETERS

num_labels = len(used_labels)
#nb_samples = len(main_database)
#nb_train_samples = math.floor(nb_samples * train_split)
#nb_valid_samples = math.floor(nb_samples * valid_split)
#nb_test_samples = nb_samples - nb_valid_samples - nb_train_samples

###############
# MAIN

test_database = load_databases(load_path)

test_x = np.array([row.reshape(64,64,1) for row in test_database[0]])
test_y = np.array([row for row in test_database[1]])

test_database = None

#!!!!! TEST ALL !!!!!
num_prediction = len(test_x)

# MAIN
if cfg_use_latest_model:
    model_files = glob.glob(path_d['Model'] + "\\model*.hdf5")
    model_load_path = max(model_files, key=os.path.getctime)
else:
    model_load_path = cfg_model_load_path

model = models.load_model(model_load_path)

if cfg_find_outliers:
    findOutliers(model, used_labels, path_d)
    sys.exit(0)

conf_mat = np.zeros((num_labels,num_labels), dtype='int')

prediction = model.predict(test_x[0:num_prediction], batch_size=batch_size, verbose=1)

for i,item in enumerate(prediction):
    pred_str = ""
    #preds = [item, used_labels]
    order = np.argsort(item)
    preds = [[used_labels[j], item[j]] for j in reversed(order)]
    for pred in preds:
        pred_str += "%.2f %s," % (pred[1], pred[0])

    pred_index = np.argmax(item)
    truth_index = np.argmax(test_y[i])
    conf_mat[truth_index][pred_index] += 1

    if cfg_print_debug:
        print("Prediction: " + pred_str)
        print("Truth: %s\n" % (used_labels[truth_index]))

    #pic_array = test_x[i]
    #img = Image.fromarray(pic_array.astype(np.uint8), 'RGB')
    #img.show()


print(conf_mat)
plot_conf_mat(conf_mat, used_labels, normal=True)

